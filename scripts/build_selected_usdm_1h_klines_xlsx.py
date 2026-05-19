from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WORKSPACE = Path(__file__).resolve().parents[1]
REPORT_DIR = WORKSPACE / "reports"
MANIFEST = REPORT_DIR / "binance_usdm_1m_kline_5symbols_latest_manifest.json"
EXCEL_MAX_ROWS = 1_048_576


DB_HEADERS_ZH = {
    "id": "DB行ID(id)",
    "symbol": "市场(symbol)",
    "interval": "周期(interval)",
    "timestamp": "开盘时间(ms)",
    "timestamp_utc": "开盘时间(UTC)",
    "open": "开盘价(open)",
    "high": "最高价(high)",
    "low": "最低价(low)",
    "close": "收盘价(close)",
    "volume": "成交量(volume)",
    "close_time": "收盘时间(ms)",
    "close_time_utc": "收盘时间(UTC)",
    "quote_volume": "成交额(quote_volume)",
    "trades": "成交笔数(trades)",
    "taker_buy_base_volume": "主动买入量(taker_buy_base_volume)",
    "taker_buy_quote_volume": "主动买入额(taker_buy_quote_volume)",
    "source": "数据来源(source)",
    "created_at": "创建时间(created_at)",
    "updated_at": "更新时间(updated_at)",
    "is_delisted": "是否下架(is_delisted)",
    "contract_type": "合约类型(contract_type)",
    "quote_asset": "计价资产(quote_asset)",
    "margin_asset": "保证金资产(margin_asset)",
}

SOURCE_HEADERS_ZH = {
    "symbol": "市场(symbol)",
    "interval": "周期(interval)",
    "timestamp": "开盘时间(ms)",
    "timestamp_utc": "开盘时间(UTC)",
    "open": "开盘价(open)",
    "high": "最高价(high)",
    "low": "最低价(low)",
    "close": "收盘价(close)",
    "volume": "成交量(volume)",
    "close_time": "收盘时间(ms)",
    "close_time_utc": "收盘时间(UTC)",
    "quote_volume": "成交额(quote_volume)",
    "trades": "成交笔数(trades)",
    "taker_buy_base_volume": "主动买入量(taker_buy_base_volume)",
    "taker_buy_quote_volume": "主动买入额(taker_buy_quote_volume)",
}

DIFF_HEADERS = ["symbol", "timestamp", "timestamp_utc", "field", "reason", "db_value", "source_value", "note"]
DIFF_HEADERS_ZH = ["市场", "开盘时间(ms)", "开盘时间(UTC)", "异常字段", "异常类型", "DB值", "源值", "异常点说明"]


def styled_cell(ws, value, *, title=False, header=False, text_format=False):
    thin = Side(style="thin", color="D9E2F3")
    cell = WriteOnlyCell(ws, value=value)
    cell.alignment = Alignment(vertical="top", wrap_text=False)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if title:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, color="000000", size=13)
    if header:
        cell.fill = PatternFill("solid", fgColor="BDD7EE")
        cell.font = Font(bold=True, color="000000")
    if text_format:
        cell.number_format = "@"
        cell.quotePrefix = True
    return cell


def csv_counter(path: Path, field: str) -> Counter:
    counts = Counter()
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            counts[row.get(field, "")] += 1
    return counts


def csv_data_row_count(path: Path) -> int:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return max(sum(1 for _ in f) - 1, 0)


def append_file_index_sheet(wb, sheet_name: str, title: str, csv_path: Path, row_count: int):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 96
    ws.append([styled_cell(ws, title, title=True), styled_cell(ws, "文件过大，未写入 Excel 明细", title=True)])
    ws.append([styled_cell(ws, "项目", header=True), styled_cell(ws, "值", header=True)])
    for key, value in [
        ("CSV路径", str(csv_path)),
        ("数据行数", str(row_count)),
        ("说明", f"Excel 单个 sheet 最大 {EXCEL_MAX_ROWS} 行；该原数据请直接查看 CSV。"),
    ]:
        ws.append([styled_cell(ws, key), styled_cell(ws, value, text_format=True)])
    return ws


def append_csv_sheet(wb, sheet_name: str, title: str, csv_path: Path, header_map: dict[str, str], widths: list[int]):
    row_count = csv_data_row_count(csv_path)
    if row_count + 2 > EXCEL_MAX_ROWS:
        return append_file_index_sheet(wb, sheet_name, title, csv_path, row_count)

    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        for idx, width in enumerate(widths[: len(fields)], 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.append([styled_cell(ws, title, title=True), styled_cell(ws, f"来源文件：{csv_path.name}", title=True)])
        ws.append([styled_cell(ws, header_map.get(field, field), header=True) for field in fields])
        for row in reader:
            ws.append([styled_cell(ws, row.get(field, ""), text_format=True) for field in fields])
    return ws


def main() -> int:
    meta = json.loads(MANIFEST.read_text(encoding="utf-8"))
    diff_csv = Path(meta["files"]["differences_csv"])
    db_csv = Path(meta["files"]["db_raw_csv"])
    source_csv = Path(meta["files"]["source_raw_csv"])
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    by_field = csv_counter(diff_csv, "field")
    by_reason = csv_counter(diff_csv, "reason")
    diff_row_count = csv_data_row_count(diff_csv)

    wb = Workbook(write_only=True)

    summary = wb.create_sheet("汇总")
    summary.freeze_panes = "A3"
    for idx, width in enumerate([34, 22, 24, 26, 26, 26, 26, 54], 1):
        summary.column_dimensions[get_column_letter(idx)].width = width
    summary.append([
        styled_cell(summary, f"Binance USD-M 合约 {meta['interval']} K线专项对比", title=True),
        styled_cell(summary, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", title=True),
    ])
    summary.append([styled_cell(summary, "参数", header=True), styled_cell(summary, "值", header=True)])
    for key, value in [
        ("DB表", meta["table"]),
        ("市场", ", ".join(meta["symbols"])),
        ("周期", meta["interval"]),
        ("开始时间", meta["start_utc"]),
        ("结束时间", meta["end_utc"]),
        ("结束说明", meta["range_note"]),
        ("DB原数据行数", meta["db_rows"]),
        ("源原数据行数", meta["source_rows"]),
        ("差异总数", meta["differences"]),
    ]:
        summary.append([styled_cell(summary, key), styled_cell(summary, str(value), text_format=True)])
    summary.append([styled_cell(summary, ""), styled_cell(summary, "")])
    symbol_headers = ["市场", "DB行数", "源行数", "差异数", "DB缺行", "源缺行", "字段不一致", "主要不一致字段", "DB最大时间", "源最大时间"]
    summary.append([styled_cell(summary, h, header=True) for h in symbol_headers])
    for row in meta["summary_by_symbol"]:
        summary.append([
            styled_cell(summary, row["symbol"], text_format=True),
            styled_cell(summary, str(row["db_rows"]), text_format=True),
            styled_cell(summary, str(row["source_rows"]), text_format=True),
            styled_cell(summary, str(row["differences"]), text_format=True),
            styled_cell(summary, str(row["missing_db_rows"]), text_format=True),
            styled_cell(summary, str(row["missing_source_rows"]), text_format=True),
            styled_cell(summary, str(row["value_mismatches"]), text_format=True),
            styled_cell(summary, row["top_mismatch_fields"], text_format=True),
            styled_cell(summary, row["db_max_time"], text_format=True),
            styled_cell(summary, row["source_max_time"], text_format=True),
        ])
    summary.append([styled_cell(summary, ""), styled_cell(summary, "")])
    summary.append([styled_cell(summary, "按异常字段", header=True), styled_cell(summary, "数量", header=True)])
    for field, count in by_field.most_common():
        summary.append([styled_cell(summary, field, text_format=True), styled_cell(summary, str(count), text_format=True)])
    summary.append([styled_cell(summary, ""), styled_cell(summary, "")])
    summary.append([styled_cell(summary, "按异常类型", header=True), styled_cell(summary, "数量", header=True)])
    for reason, count in by_reason.most_common():
        summary.append([styled_cell(summary, reason, text_format=True), styled_cell(summary, str(count), text_format=True)])

    if diff_row_count + 2 > EXCEL_MAX_ROWS:
        append_file_index_sheet(wb, "对比结果", "字段级差异明细", diff_csv, diff_row_count)
    else:
        diff = wb.create_sheet("对比结果")
        diff.freeze_panes = "A3"
        for idx, width in enumerate([16, 22, 24, 24, 22, 34, 34, 44], 1):
            diff.column_dimensions[get_column_letter(idx)].width = width
        diff.append([styled_cell(diff, "字段级差异明细", title=True), styled_cell(diff, f"差异总数：{meta['differences']}", title=True)])
        diff.append([styled_cell(diff, h, header=True) for h in DIFF_HEADERS_ZH])
        with diff_csv.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                diff.append([styled_cell(diff, row.get(field, ""), text_format=True) for field in DIFF_HEADERS])

    append_csv_sheet(
        wb,
        "DB原数据",
        "DB 原数据",
        db_csv,
        DB_HEADERS_ZH,
        [18, 16, 14, 22, 24, 18, 18, 18, 18, 24, 22, 24, 24, 18, 28, 28, 18, 24, 24, 18, 18, 18, 18],
    )
    append_csv_sheet(
        wb,
        "源原数据",
        "Binance 源原数据",
        source_csv,
        SOURCE_HEADERS_ZH,
        [16, 14, 22, 24, 18, 18, 18, 18, 24, 22, 24, 24, 18, 28, 28],
    )

    latest_xlsx = REPORT_DIR / "binance_usdm_1m_kline_5symbols_20240101_to_now_latest_zh.xlsx"
    stamped_xlsx = REPORT_DIR / f"binance_usdm_1m_kline_5symbols_20240101_to_now_{stamp}_zh.xlsx"
    wb.save(latest_xlsx)
    stamped_xlsx.write_bytes(latest_xlsx.read_bytes())

    check = load_workbook(latest_xlsx, read_only=True, data_only=False)
    ws = check["对比结果"]
    print("xlsx_latest=", latest_xlsx)
    print("xlsx_stamped=", stamped_xlsx)
    print("A1=", ws["A1"].value)
    print("A2:H2=", [ws.cell(2, c).value for c in range(1, 9)])
    if meta["differences"]:
        print("G3=", ws["G3"].value, "type=", ws["G3"].data_type)
    print("db_rows=", meta["db_rows"], "source_rows=", meta["source_rows"], "differences=", meta["differences"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
